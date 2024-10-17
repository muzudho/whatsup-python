#
# cd using_openpyxl
# python let_s_make_tree_view_on_excel.py
#
# エクセルで樹形図を描こう
#

import traceback
import datetime
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side

from library.database import TreeNode, TreeRecord, TreeTable
from library.views.tree import TreeView

CSV_FILE_PATH = '../data/tree_shiritori.csv'
WB_FILE_PATH = '../temp/tree.xlsx'


class TreeDrawer():
    """エクセルで罫線などを駆使して、樹形図を描画します"""


    def __init__(self, df, wb):
        self._df = df
        self._wb = wb
        self._prev_record = TreeRecord.new_empty()
        self._curr_record = TreeRecord.new_empty()
        self._next_record = TreeRecord.new_empty()


    def forward_cursor(self, next_record):
        """送り出し

        Parameters
        ----------
        next_record : Record
            次行
        """
        self._prev_record = self._curr_record
        self._curr_record = self._next_record
        self._next_record = next_record


    def on_header(self):

        # 変数名の短縮
        ws = self._wb['Tree']


        # 列の幅設定
        # width はだいたい 'ＭＳ Ｐゴシック' サイズ11 の半角英文字の個数

        ws.column_dimensions['A'].width = 4     # no
        ws.column_dimensions['B'].width = 6     # "葉"
        ws.column_dimensions['C'].width = 20    # 根
        ws.column_dimensions['D'].width = 2     # 第１層　親側辺
        ws.column_dimensions['E'].width = 4     #       　子側辺
        ws.column_dimensions['F'].width = 20    #         節
        ws.column_dimensions['G'].width = 2     # 第２層  親側辺
        ws.column_dimensions['H'].width = 4     #         子側辺
        ws.column_dimensions['I'].width = 20    #         節
        ws.column_dimensions['J'].width = 2     # 第３層  親側辺
        ws.column_dimensions['K'].width = 4     #         子側辺
        ws.column_dimensions['L'].width = 20    #         節
        ws.column_dimensions['M'].width = 2     # 第４層  親側辺
        ws.column_dimensions['N'].width = 4     #         子側辺
        ws.column_dimensions['O'].width = 20    #         節


        # 行の高さ設定
        # height の単位はポイント。初期値 8。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
        ws.row_dimensions[1].height = 13
        ws.row_dimensions[2].height = 13


        # 第１行
        # ------
        # ヘッダー行にする
        row_th = 1

        # NOTE データテーブルではなく、ビュー用途なので、テーブルとしての機能性は無視しています
        # A の代わりに {xl.utils.get_column_letter(1)} とも書ける
        ws[f'A{row_th}'] = 'No'
        # 第2列は空
        ws[f'C{row_th}'] = '├─根─┤'
        ws[f'D{row_th}'] = '├'
        ws[f'E{row_th}'] = '─'
        ws[f'F{row_th}'] = '第1層─┤'
        ws[f'G{row_th}'] = '├'
        ws[f'H{row_th}'] = '─'
        ws[f'I{row_th}'] = '第2層─┤'
        ws[f'J{row_th}'] = '├'
        ws[f'K{row_th}'] = '─'
        ws[f'L{row_th}'] = '第3層─┤'
        ws[f'M{row_th}'] = '├'
        ws[f'N{row_th}'] = '─'
        ws[f'O{row_th}'] = '第4層─┤'

        # 第２行
        # ------
        # 空行にする
        row_th = 2


    def on_each_record(self, next_row_number, next_record):
        """先読みで最初の１回を空振りさせるので、２件目から本処理です"""

        # 事前送り出し
        self.forward_cursor(next_record=next_record)

        if self._curr_record.no is None:
            print(f"[{datetime.datetime.now()}] 第{self._curr_record.no}葉 現在レコードのnoがナンだから無視（先読みのため、初回は空回し）")
            pass


        else:
            # 変数名短縮
            ws = self._wb['Tree']


            # ３行目～６行目
            # --------------
            # データは３行目から、１かたまり３行を使って描画する
            curr_row_number = next_row_number - 1
            row1_th = curr_row_number * 3 + 3
            row2_th = curr_row_number * 3 + 3 + 1
            row3_th = curr_row_number * 3 + 3 + 2
            three_row_numbers = [row1_th, row2_th, row3_th]

            # 行の高さ設定
            # height の単位はポイント。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
            ws.row_dimensions[row1_th].height = 13
            ws.row_dimensions[row2_th].height = 13
            ws.row_dimensions[row3_th].height = 6

            ws[f'A{row1_th}'].value = self._curr_record.no
            ws[f'B{row1_th}'].value = '葉'


            def draw_edge(depth_th, three_column_names, three_row_numbers):
                """
                Parameters
                ----------
                depth_th : int
                    第何層。根層は 0
                """

                # 色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)

                # 罫線
                #
                #   style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
                #
                BLACK = '000000'
                side = Side(style='thick', color=BLACK)

                # DEBUG_POINT
                # 罫線に色を付けると、デバッグしやすいです
                if True:
                    red_side = Side(style='thick', color=BLACK)
                    orange_side = Side(style='thick', color=BLACK)
                    green_side = Side(style='thick', color=BLACK)
                    blue_side = Side(style='thick', color=BLACK)
                    cyan_side = Side(style='thick', color=BLACK)
                    magenta_side = Side(style='thick', color=BLACK)
                else:
                    red_side = Side(style='thick', color='FF0000')
                    orange_side = Side(style='thick', color='FFCC00')
                    green_side = Side(style='thick', color='00FF00')
                    blue_side = Side(style='thick', color='0000FF')
                    cyan_side = Side(style='thick', color='00FFFF')
                    magenta_side = Side(style='thick', color='FF00FF')

                # ─字接続は赤
                border_to_parent_horizontal = Border(bottom=red_side)
                under_border_to_child_horizontal = Border(bottom=red_side)
                # │字接続は緑
                leftside_border_to_vertical = Border(left=green_side)
                # ┬字接続は青
                border_to_parent_downward = Border(bottom=blue_side)
                under_border_to_child_downward = Border(bottom=blue_side)
                leftside_border_to_child_downward = Border(left=blue_side)
                # ├字接続は青緑
                l_letter_border_to_child_rightward = Border(left=cyan_side, bottom=cyan_side)
                leftside_border_to_child_rightward = Border(left=cyan_side)
                # └字接続は橙
                l_letter_border_to_child_upward = Border(left=orange_side, bottom=orange_side)


                prerow_nd = self._prev_record.node_at(depth_th=depth_th)
                nd = self._curr_record.node_at(depth_th=depth_th)

                if nd is None:
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}葉 第{depth_th}層  nd がナンのノードは無視")
                    return

                # nd.text が NaN のノードは無視
                elif pd.isnull(nd.text):
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}葉 第{depth_th}層  空欄")
                    return


                cn1 = three_column_names[0]
                cn2 = three_column_names[1]
                cn3 = three_column_names[2]
                row1_th = three_row_numbers[0]
                row2_th = three_row_numbers[1]
                row3_th = three_row_numbers[2]


                # 先祖から自分までが同じノードテキストのレコードが続くなら
                if TreeView.is_same_between_ancestor_and_myself_as_avobe(
                        curr_record=self._curr_record,
                        prev_record=self._prev_record,
                        depth_th=depth_th):

                    # 垂直線
                    #
                    #   |    leftside_border
                    # ..+..  
                    #   |    leftside_border
                    #   |    leftside_border
                    #
                    if TreeView.is_same_between_ancestor_and_myself_as_avobe(
                            curr_record=self._curr_record,
                            prev_record=self._prev_record,
                            depth_th=depth_th):
                        print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}葉 第{depth_th}層  │")
                        
                        ws[f'{cn2}{row1_th}'].border = leftside_border_to_vertical
                        ws[f'{cn2}{row2_th}'].border = leftside_border_to_vertical
                        ws[f'{cn2}{row3_th}'].border = leftside_border_to_vertical
                    
                    else:
                        print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}葉 第{depth_th}層  空欄")
                        pass

                    return
                

                # # １列目：親ノードから伸びてきた枝
                # #
                # #   .
                # # --...
                # #   .
                # #
                # # 前ラウンドにノードがあれば、接続線を引く
                # #
                # if TreeView.can_connect_to_parent(
                #         curr_record=self._curr_record,
                #         prev_record=self._prev_record,
                #         depth_th=depth_th):
                #     ws[f'{cn1}{row1_th}'].border = border_to_parent


                # ２列目：エッジ・テキスト
                ws[f'{cn2}{row1_th}'].value = nd.edge_text


                # 子ノードへの接続は４種類の線がある
                #
                # (1) ─字
                #   .    under_border
                # ...__  
                #   .    None
                #   .    None
                #
                # (2) ┬字
                #   .    under_border
                # ..+__  
                #   |    leftside_border
                #   |    leftside_border
                #
                # (3) ├字
                #   |    l_letter_border
                # ..+__  
                #   |    leftside_border
                #   |    leftside_border
                #
                # (4) └字
                #   |    l_letter_border
                # ..+__  
                #   .    None
                #   .    None
                #
                kind = TreeView.get_kind_connect_to_child(
                        prev_record=self._prev_record,
                        curr_record=self._curr_record,
                        next_record=self._next_record,
                        depth_th=depth_th)

                if kind == '─字':
                    ws[f'{cn1}{row1_th}'].border = border_to_parent_horizontal
                    ws[f'{cn2}{row1_th}'].border = under_border_to_child_horizontal
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}葉 第{depth_th}層  ─ {nd.edge_text}")
                
                elif kind == '┬字':
                    ws[f'{cn1}{row1_th}'].border = border_to_parent_downward
                    ws[f'{cn2}{row1_th}'].border = under_border_to_child_downward
                    ws[f'{cn2}{row2_th}'].border = leftside_border_to_child_downward
                    ws[f'{cn2}{row3_th}'].border = leftside_border_to_child_downward
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}葉 第{depth_th}層  ┬ {nd.edge_text}")

                elif kind == '├字':
                    ws[f'{cn2}{row1_th}'].border = l_letter_border_to_child_rightward
                    ws[f'{cn2}{row2_th}'].border = leftside_border_to_child_rightward
                    ws[f'{cn2}{row3_th}'].border = leftside_border_to_child_rightward
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}葉 第{depth_th}層  ├ {nd.edge_text}")

                elif kind == '└字':
                    ws[f'{cn2}{row1_th}'].border = l_letter_border_to_child_upward
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}葉 第{depth_th}層  └ {nd.edge_text}")
                
                else:
                    raise ValueError(f"{kind=}")


            def draw_node(depth_th, three_column_names, three_row_numbers):
                """節を描きます

                Parameters
                ----------
                node : TreeNode
                    節
                depth_th : int
                    第何層。根層は 0
                """

                prerow_nd = self._prev_record.node_at(depth_th=depth_th)
                nd = self._curr_record.node_at(depth_th=depth_th)

                if nd is None:
                    #print(f"[{datetime.datetime.now()}] 鉛筆(節) 第{self._curr_record.no}葉 第{depth_th}層  nd がナンのノードは無視")
                    return

                elif pd.isnull(nd.text):
                    #print(f"[{datetime.datetime.now()}] 鉛筆(節) 第{self._curr_record.no}葉 第{depth_th}層  nd.text が NaN のノードは無視")
                    return

                # 先祖から自分までが同じノードテキストのレコードが続くなら省く
                elif TreeView.is_same_between_ancestor_and_myself_as_avobe(
                        curr_record=self._curr_record,
                        prev_record=self._prev_record,
                        depth_th=depth_th):
                    #print(f"[{datetime.datetime.now()}] 鉛筆(節) 第{self._curr_record.no}葉 第{depth_th}層  同じディレクトリーは描画を省く")
                    return


                cn3 = three_column_names[2]
                row1_th = three_row_numbers[0]
                row2_th = three_row_numbers[1]
                row3_th = three_row_numbers[2]

                # 色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)
                node_bgcolor = PatternFill(patternType='solid', fgColor='FFFFCC')

                # 罫線
                #
                #   style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
                #
                side = Side(style='thick', color='000000')
                upside_node_border = Border(top=side, left=side, right=side)
                downside_node_border = Border(bottom=side, left=side, right=side)

                print(f"[{datetime.datetime.now()}] 鉛筆(節) 第{self._curr_record.no}葉 第{depth_th}層  □ {nd.text}")
                ws[f'{cn3}{row1_th}'].value = nd.text
                ws[f'{cn3}{row1_th}'].fill = node_bgcolor
                ws[f'{cn3}{row1_th}'].border = upside_node_border
                ws[f'{cn3}{row2_th}'].fill = node_bgcolor
                ws[f'{cn3}{row2_th}'].border = downside_node_border


            # 第０層
            # ------
            draw_node(depth_th=0, three_column_names=[None, None, 'C'], three_row_numbers=three_row_numbers)


            # 第１層
            # ------
            depth_th = 1
            three_column_names=['D', 'E', 'F']
            draw_edge(depth_th=depth_th, three_column_names=three_column_names, three_row_numbers=three_row_numbers)
            draw_node(depth_th=depth_th, three_column_names=three_column_names, three_row_numbers=three_row_numbers)


            # 第２層
            # ------
            depth_th = 2
            three_column_names=['G', 'H', 'I']
            draw_edge(depth_th=depth_th, three_column_names=three_column_names, three_row_numbers=three_row_numbers)
            draw_node(depth_th=depth_th, three_column_names=three_column_names, three_row_numbers=three_row_numbers)


            # 第３層
            # ------
            depth_th = 3
            three_column_names=['J', 'K', 'L']
            draw_edge(depth_th=depth_th, three_column_names=three_column_names, three_row_numbers=three_row_numbers)
            draw_node(depth_th=depth_th, three_column_names=three_column_names, three_row_numbers=three_row_numbers)


            # 第４層
            # ------
            depth_th = 4
            three_column_names=['M', 'N', 'O']
            draw_edge(depth_th=depth_th, three_column_names=three_column_names, three_row_numbers=three_row_numbers)
            draw_node(depth_th=depth_th, three_column_names=three_column_names, three_row_numbers=three_row_numbers)


class TreeEraser():
    """要らない罫線を消す"""


    def __init__(self, wb):
        self._wb = wb


    def erase_unnecessary_border_by_column(self, column_alphabet):
        """不要な境界線を消す"""

        # DEBUG_POINT
        # デバッグ時は、罫線を消すのではなく、灰色に変えると見やすいです
        if True:
            # 罫線無し
            striked_border = None
        else:
            # 色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)
            #
            # 罫線
            #
            #   style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
            #
            # 見え消し用（デバッグに使う）
            striked_side = Side(style='thick', color='DDDDDD')
            # 見え消し用の罫線
            striked_border = Border(left=striked_side)


        # 変数名の短縮
        ws = self._wb['Tree']


        # 最後に見つけた、セルの左辺に罫線がなく、下辺に太い罫線がある行をリセット
        row_th_of_prev_last_underline = -1
        row_th_of_last_underline = -1


        # 第3行から
        row_th = 3
        while row_th <= ws.max_row: # 最終行まで全部見る

            while True: # 仕切り直しの１セット
                shall_break = False

                # 罫線を確認
                #
                #   .
                # ..+--  下向きの罫線が最後に出た箇所を調べる
                #   |
                #
                border = ws[f'{column_alphabet}{row_th}'].border
                if border is not None:
                    #print(f"[{datetime.datetime.now()}] 消しゴム {column_alphabet}列第{row_th}行 境界線有り {border=}")

                    # セルの左辺に太い罫線が引かれており...
                    if border.left is not None and border.left.style == 'thick':
                        # セルの下辺にも太い罫線が引かれていれば、［ラスト・シブリング］だ
                        if border.bottom is not None and border.bottom.style == 'thick':
                            row_th_of_prev_last_underline = -1
                            row_th_of_last_underline = -1
                            print(f"[{datetime.datetime.now()}] 消しゴム {column_alphabet}列第{row_th}行 ラスト・シブリングなので、最後に見つけた左辺に罫線のないアンダーラインのことは忘れて仕切り直し")
                            shall_break = True

                        # 次行へ読み進めていく
                        else:
                            print(f"[{datetime.datetime.now()}] 消しゴム {column_alphabet}列第{row_th}行 左側に罫線")
                            pass

                    # セルの左辺に太い罫線が引かれておらず、セルの下辺に太い罫線が引かれていたら、つながっていない垂線だ。それが第何行か覚えておいて仕切り直す
                    elif border.bottom is not None and border.bottom.style == 'thick':
                        row_th_of_prev_last_underline = row_th_of_last_underline
                        row_th_of_last_underline = row_th
                        print(f"[{datetime.datetime.now()}] 消しゴム {column_alphabet}列第{row_th}行 最後に見つけた、左辺に罫線のないアンダーラインが第何行か覚えておく（第{row_th_of_last_underline}行）（１つ前は第{row_th_of_prev_last_underline}行）")
                        shall_break = True

                    # セルの左辺にも、下辺にも、太い罫線が引かれていなければ、仕切り直し
                    else:
                        shall_break = True
                        print(f"[{datetime.datetime.now()}] 消しゴム {column_alphabet}列第{row_th}行 セルの左辺にも下辺にも罫線が引かれていなかったので、仕切り直し")


                row_th += 1

                if shall_break:
                    break


            # 消しゴムを掛ける
            start_row_to_erase = row_th_of_prev_last_underline + 1
            end_row_to_erase = row_th_of_last_underline
            #print(f"[{datetime.datetime.now()}] 消しゴム {column_alphabet}列第{row_th}行 仕切り直し {row_th_of_last_underline=} {start_row_to_erase=} {end_row_to_erase=}")

            if row_th_of_last_underline != -1 and 0 < start_row_to_erase and start_row_to_erase < end_row_to_erase:
                print(f"[{datetime.datetime.now()}] 消しゴム {column_alphabet}列 消しゴムを掛けたいのは第{start_row_to_erase}～{end_row_to_erase - 1}行")
                for row_th_to_erase in range(start_row_to_erase, end_row_to_erase):
                    # 消すか、見え消しにするか切り替えられるようにしておく
                    ws[f'{column_alphabet}{row_th_to_erase}'].border = striked_border

        print(f"[{datetime.datetime.now()}] 消しゴム {column_alphabet}列第{row_th}行 消しゴム掛け終わり（最終は第{ws.max_row}行）")


    def execute(self):

        # TODO 可変長に対応したい
        # 指定の列の左側の垂直の罫線を見ていく
        self.erase_unnecessary_border_by_column(column_alphabet='E')
        self.erase_unnecessary_border_by_column(column_alphabet='H')
        self.erase_unnecessary_border_by_column(column_alphabet='K')
        self.erase_unnecessary_border_by_column(column_alphabet='N')


########################################
# コマンドから実行時
########################################
if __name__ == '__main__':
    """コマンドから実行時"""

    try:
        # ワークブックを生成
        wb = xl.Workbook()

        # Tree シートを作成
        wb.create_sheet('Tree')

        # 既存の Sheet シートを削除
        wb.remove(wb['Sheet'])

        # CSV読込
        tree_table = TreeTable.from_csv(file_path=CSV_FILE_PATH)

        # CSV確認
        print(f"""\
tree_table.df:
{tree_table.df}""")

        tree_drawer = TreeDrawer(df=tree_table.df, wb=wb)

        # GTWB の Sheet シートへのヘッダー書出し
        tree_drawer.on_header()

        # GTWB の Sheet シートへの各行書出し
        tree_table.for_each(on_each=tree_drawer.on_each_record)

        # 最終行の実行
        tree_drawer.on_each_record(next_row_number=len(tree_table.df), next_record=TreeRecord.new_empty())


        # DEBUG_POINT
        if True:
            # 要らない罫線を消す
            #
            #   このコードを不活性にして、必要な線は全部描かれていることを確認してください
            #
            tree_eraser = TreeEraser(wb=wb)
            tree_eraser.execute()
        else:
            print(f"消しゴム　使用中止中")


        # ワークブックの保存
        wb.save(WB_FILE_PATH)


    except Exception as err:
        print(f"[unexpected error] {err=}  {type(err)=}")

        # スタックトレース表示
        print(traceback.format_exc())
