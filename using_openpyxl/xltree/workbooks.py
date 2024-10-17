import datetime
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side

from xltree.database import TreeNode, TreeRecord
from xltree.models import TreeModel


class TreeDrawer():
    """エクセルで罫線などを駆使して、樹形図を描画します"""


    def __init__(self, tree_table, ws, config):
        """初期化
        
        Parameters
        ----------
        df : DataFrame
            データフレーム
        ws : openpyxl.Worksheet
            ワークシート
        """
        self._tree_table = tree_table
        self._ws = ws
        self._config = config

        self._prev_record = TreeRecord.new_empty(specified_length_of_nodes=self._tree_table.actual_length_of_nodes)
        self._curr_record = TreeRecord.new_empty(specified_length_of_nodes=self._tree_table.actual_length_of_nodes)
        self._next_record = TreeRecord.new_empty(specified_length_of_nodes=self._tree_table.actual_length_of_nodes)


    def render(self):
        """描画"""

        # 対象シートへ列ヘッダー書出し
        self._on_header()

        # 対象シートへの各行書出し
        self._tree_table.for_each(on_each=self._on_each_record)

        # 最終行の実行
        self._on_each_record(next_row_number=len(self._tree_table.df), next_record=TreeRecord.new_empty(specified_length_of_nodes=self._tree_table.actual_length_of_nodes))


    def _forward_cursor(self, next_record):
        """送り出し

        Parameters
        ----------
        next_record : Record
            次行
        """
        self._prev_record = self._curr_record
        self._curr_record = self._next_record
        self._next_record = next_record


    def _on_header(self):

        # 変数名の短縮
        ws = self._ws


        # 列の幅設定
        column_width_dict = {}
        column_width_dict['A'] = self._config.dictionary['no_width']                        # no
        column_width_dict['B'] = self._config.dictionary['row_header_separator_width']      # 空列
        column_width_dict['C'] = self._config.dictionary['node_width']                      # 根

        head_column_th = 4
        for node_th in range(1, self._tree_table.actual_length_of_nodes):
            column_width_dict[xl.utils.get_column_letter(head_column_th    )] = self._config.dictionary['parent_side_edge_width']   # 第n層  親側辺
            column_width_dict[xl.utils.get_column_letter(head_column_th + 1)] = self._config.dictionary['child_side_edge_width']    #        子側辺
            column_width_dict[xl.utils.get_column_letter(head_column_th + 2)] = self._config.dictionary['node_width']               #        節
            head_column_th += 3


        for name, width in column_width_dict.items():
            ws.column_dimensions[name].width = width


        # 行の高さ設定
        # height の単位はポイント。初期値 8。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
        row_height_dict = {
            1: self._config.dictionary['header_height'],
            2: self._config.dictionary['column_header_separator_height'],
        }

        for row_number, height in row_height_dict.items():
            ws.row_dimensions[row_number].height = height


        # 第１行
        # ------
        # ヘッダー行にする
        row_th = 1

        # TODO 可変長ノード数への対応
        # NOTE データテーブルではなく、ビュー用途なので、テーブルとしての機能性は無視しています
        # A の代わりに {xl.utils.get_column_letter(1)} とも書ける
        ws[f'A{row_th}'] = 'No'
        # 第2列は空
        ws[f'C{row_th}'] = 'Root'
        ws[f'F{row_th}'] = '1st'
        ws[f'I{row_th}'] = '2nd'
        ws[f'L{row_th}'] = '3rd'
        ws[f'O{row_th}'] = '4th'


        # 背景色
        #
        #   色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)
        #
        fgcolor_list = [
            Font(color='EEEEEE'),
            Font(color='111111'),
        ]
        bgcolor_list = [
            PatternFill(patternType='solid', fgColor='333333'),
            PatternFill(patternType='solid', fgColor='CCCCCC')]

        ws[f'A{row_th}'].fill = bgcolor_list[1]
        ws[f'A{row_th}'].font = fgcolor_list[1]
        ws[f'B{row_th}'].fill = bgcolor_list[1]

        ws[f'C{row_th}'].fill = bgcolor_list[0]
        ws[f'C{row_th}'].font = fgcolor_list[0]

        ws[f'D{row_th}'].fill = bgcolor_list[1]
        ws[f'E{row_th}'].fill = bgcolor_list[1]
        ws[f'F{row_th}'].fill = bgcolor_list[1]
        ws[f'F{row_th}'].font = fgcolor_list[1]

        ws[f'G{row_th}'].fill = bgcolor_list[0]
        ws[f'H{row_th}'].fill = bgcolor_list[0]
        ws[f'I{row_th}'].fill = bgcolor_list[0]
        ws[f'I{row_th}'].font = fgcolor_list[0]

        ws[f'J{row_th}'].fill = bgcolor_list[1]
        ws[f'K{row_th}'].fill = bgcolor_list[1]
        ws[f'L{row_th}'].fill = bgcolor_list[1]
        ws[f'L{row_th}'].font = fgcolor_list[1]

        ws[f'M{row_th}'].fill = bgcolor_list[0]
        ws[f'N{row_th}'].fill = bgcolor_list[0]
        ws[f'O{row_th}'].fill = bgcolor_list[0]
        ws[f'O{row_th}'].font = fgcolor_list[0]


        # 第２行
        # ------
        # 空行にする
        row_th = 2


    def _on_each_record(self, next_row_number, next_record):
        """先読みで最初の１回を空振りさせるので、２件目から本処理です"""

        # 事前送り出し
        self._forward_cursor(next_record=next_record)

        if self._curr_record.no is None:
            print(f"[{datetime.datetime.now()}] 第{self._curr_record.no}件 最初のレコードは先読みのため、空回しします")
            pass


        else:
            # 変数名短縮
            ws = self._ws


            # ３行目～６行目
            # --------------
            # データは３行目から、１かたまり３行を使って描画する
            HEADER_HEIGHT = 3
            RECORD_HEIGHT = 3
            curr_row_number = next_row_number - 1
            row1_th = curr_row_number * RECORD_HEIGHT + HEADER_HEIGHT
            row2_th = curr_row_number * RECORD_HEIGHT + HEADER_HEIGHT + 1
            row3_th = curr_row_number * RECORD_HEIGHT + HEADER_HEIGHT + 2
            three_row_numbers = [row1_th, row2_th, row3_th]

            # 行の高さ設定
            # height の単位はポイント。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
            ws.row_dimensions[row1_th].height = 13
            ws.row_dimensions[row2_th].height = 13
            ws.row_dimensions[row3_th].height = 6

            ws[f'A{row1_th}'].value = self._curr_record.no
            # B列は空


            def draw_edge(depth_th, three_column_names, three_row_numbers):
                """
                Parameters
                ----------
                depth_th : int
                    第何層。根層は 0
                """

                # 罫線
                #
                #   style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
                #   色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)
                #
                BLACK = '000000'
                side = Side(style='thick', color=BLACK)

                # DEBUG_TIPS: 罫線に色を付けると、デバッグしやすいです
                if True:
                    red_side = Side(style='thick', color=BLACK)
                    orange_side = Side(style='thick', color=BLACK)
                    green_side = Side(style='thick', color=BLACK)
                    blue_side = Side(style='thick', color=BLACK)
                    cyan_side = Side(style='thick', color=BLACK)
                else:
                    red_side = Side(style='thick', color='FF0000')
                    orange_side = Side(style='thick', color='FFCC00')
                    green_side = Side(style='thick', color='00FF00')
                    blue_side = Side(style='thick', color='0000FF')
                    cyan_side = Side(style='thick', color='00FFFF')

                # ─字  赤
                border_to_parent_horizontal = Border(bottom=red_side)
                under_border_to_child_horizontal = Border(bottom=red_side)
                # │字  緑
                leftside_border_to_vertical = Border(left=green_side)
                # ┬字  青
                border_to_parent_downward = Border(bottom=blue_side)
                under_border_to_child_downward = Border(bottom=blue_side)
                leftside_border_to_child_downward = Border(left=blue_side)
                # ├字  青緑
                l_letter_border_to_child_rightward = Border(left=cyan_side, bottom=cyan_side)
                leftside_border_to_child_rightward = Border(left=cyan_side)
                # └字  橙
                l_letter_border_to_child_upward = Border(left=orange_side, bottom=orange_side)


                nd = self._curr_record.node_at(depth_th=depth_th)

                if nd is None or pd.isnull(nd.text):
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}件 第{depth_th}層  空欄")
                    return


                cn1 = three_column_names[0]
                cn2 = three_column_names[1]
                cn3 = three_column_names[2]
                row1_th = three_row_numbers[0]
                row2_th = three_row_numbers[1]
                row3_th = three_row_numbers[2]


                # 自件と前件を比較して、根から自ノードまで、ノードテキストが等しいか？
                if TreeModel.is_same_path_as_avobe(
                        curr_record=self._curr_record,
                        prev_record=self._prev_record,
                        depth_th=depth_th):

                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}件 第{depth_th}層  │")
                    # 垂直線
                    #
                    #   |    leftside_border
                    # ..+..  
                    #   |    leftside_border
                    #   |    leftside_border
                    #                        
                    ws[f'{cn2}{row1_th}'].border = leftside_border_to_vertical
                    ws[f'{cn2}{row2_th}'].border = leftside_border_to_vertical
                    ws[f'{cn2}{row3_th}'].border = leftside_border_to_vertical
                    return


                # 子ノードへの接続は４種類の線がある
                #
                # (1) ─字
                #   .    under_border
                # ...__  
                #   .    None
                #   .    None
                #
                # (2) ┬字
                #   .    under_border
                # ..+__  
                #   |    leftside_border
                #   |    leftside_border
                #
                # (3) ├字
                #   |    l_letter_border
                # ..+__  
                #   |    leftside_border
                #   |    leftside_border
                #
                # (4) └字
                #   |    l_letter_border
                # ..+__  
                #   .    None
                #   .    None
                #
                kind = TreeModel.get_kind_of_edge(
                        prev_record=self._prev_record,
                        curr_record=self._curr_record,
                        next_record=self._next_record,
                        depth_th=depth_th)

                if kind == '─字':
                    ws[f'{cn1}{row1_th}'].border = border_to_parent_horizontal
                    ws[f'{cn2}{row1_th}'].border = under_border_to_child_horizontal
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}件 第{depth_th}層  ─ {nd.edge_text}")
                
                elif kind == '┬字':
                    ws[f'{cn1}{row1_th}'].border = border_to_parent_downward
                    ws[f'{cn2}{row1_th}'].border = under_border_to_child_downward
                    ws[f'{cn2}{row2_th}'].border = leftside_border_to_child_downward
                    ws[f'{cn2}{row3_th}'].border = leftside_border_to_child_downward
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}件 第{depth_th}層  ┬ {nd.edge_text}")

                elif kind == '├字':
                    ws[f'{cn2}{row1_th}'].border = l_letter_border_to_child_rightward
                    ws[f'{cn2}{row2_th}'].border = leftside_border_to_child_rightward
                    ws[f'{cn2}{row3_th}'].border = leftside_border_to_child_rightward
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}件 第{depth_th}層  ├ {nd.edge_text}")

                elif kind == '└字':
                    ws[f'{cn2}{row1_th}'].border = l_letter_border_to_child_upward
                    print(f"[{datetime.datetime.now()}] 鉛筆(辺) 第{self._curr_record.no}件 第{depth_th}層  └ {nd.edge_text}")
                
                else:
                    raise ValueError(f"{kind=}")
                

                # ２列目：エッジ・テキスト
                ws[f'{cn2}{row1_th}'].value = nd.edge_text


            def draw_node(depth_th, three_column_names, three_row_numbers):
                """節を描きます

                Parameters
                ----------
                node : TreeNode
                    節
                depth_th : int
                    第何層。根層は 0
                """

                nd = self._curr_record.node_at(depth_th=depth_th)

                if nd is None or pd.isnull(nd.text) or TreeModel.is_same_path_as_avobe(
                        curr_record=self._curr_record,
                        prev_record=self._prev_record,
                        depth_th=depth_th):
                    print(f"[{datetime.datetime.now()}] 鉛筆(節) 第{self._curr_record.no}件 第{depth_th}層  空欄")
                    return


                cn3 = three_column_names[2]
                row1_th = three_row_numbers[0]
                row2_th = three_row_numbers[1]
                row3_th = three_row_numbers[2]

                # 背景色
                #
                #   色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)
                #
                node_bgcolor = PatternFill(patternType='solid', fgColor='FFFFCC')

                # 罫線、背景色
                #
                #   style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
                #
                side = Side(style='thick', color='000000')
                upside_node_border = Border(top=side, left=side, right=side)
                downside_node_border = Border(bottom=side, left=side, right=side)

                print(f"[{datetime.datetime.now()}] 鉛筆(節) 第{self._curr_record.no}件 第{depth_th}層  □ {nd.text}")
                ws[f'{cn3}{row1_th}'].value = nd.text
                ws[f'{cn3}{row1_th}'].fill = node_bgcolor
                ws[f'{cn3}{row1_th}'].border = upside_node_border
                ws[f'{cn3}{row2_th}'].fill = node_bgcolor
                ws[f'{cn3}{row2_th}'].border = downside_node_border


            # 第０層
            # ------
            depth_th = 0
            if depth_th < self._tree_table.actual_length_of_nodes:
                column_letter = xl.utils.get_column_letter(3)   # 'C'
                draw_node(depth_th=depth_th, three_column_names=[None, None, column_letter], three_row_numbers=three_row_numbers)


            COLUMN_WIDTH = 3

            # 第１～最終層
            # ------------
            for depth_th in range(1, self._tree_table.actual_length_of_nodes):
                head_column_th = depth_th * COLUMN_WIDTH + 1
                if depth_th < self._tree_table.actual_length_of_nodes:
                    # 第1層は 'D', 'E', 'F'、以降、後ろにずれていく
                    column_letter_list = [
                        xl.utils.get_column_letter(head_column_th),
                        xl.utils.get_column_letter(head_column_th + 1),
                        xl.utils.get_column_letter(head_column_th + 2),
                    ]
                    draw_edge(depth_th=depth_th, three_column_names=column_letter_list, three_row_numbers=three_row_numbers)
                    draw_node(depth_th=depth_th, three_column_names=column_letter_list, three_row_numbers=three_row_numbers)


class TreeEraser():
    """要らない罫線を消す"""


    def __init__(self, tree_table, ws):
        """初期化
        
        Parameters
        ----------
        ws : openpyxl.Worksheet
            ワークシート
        """
        self._tree_table = tree_table
        self._ws = ws


    def render(self):
        """描画"""

        # 指定の列の左側の垂直の罫線を見ていく
        column_th = 5
        for node_th in range(1, self._tree_table.actual_length_of_nodes):
            self._erase_unnecessary_border_by_column(column_letter=xl.utils.get_column_letter(column_th))
            column_th += 3


    def _erase_unnecessary_border_by_column(self, column_letter):
        """不要な境界線を消す"""

        # DEBUG_TIPS: デバッグ時は、罫線を消すのではなく、灰色に変えると見やすいです
        if True:
            # 罫線無し
            striked_border = None
        else:
            # 罫線
            #
            #   style に入るもの： 'dashDot', 'dashDotDot', 'double', 'hair', 'dotted', 'mediumDashDotDot', 'dashed', 'mediumDashed', 'slantDashDot', 'thick', 'thin', 'medium', 'mediumDashDot'
            #   色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)
            #
            # 見え消し用（デバッグに使う）
            striked_side = Side(style='thick', color='DDDDDD')
            # 見え消し用の罫線
            striked_border = Border(left=striked_side)


        # 変数名の短縮
        ws = self._ws


        # 最後に見つけた、セルの左辺に罫線がなく、下辺に太い罫線がある行をリセット
        row_th_of_prev_last_underline = -1
        row_th_of_last_underline = -1


        # 第3行から
        row_th = 3
        while row_th <= ws.max_row: # 最終行まで全部見る

            while True: # 仕切り直しの１セット
                shall_break = False

                # 罫線を確認
                #
                #   .
                # ..+--  下向きの罫線が最後に出た箇所を調べる
                #   |
                #
                border = ws[f'{column_letter}{row_th}'].border
                if border is not None:
                    # セルの左辺に太い罫線が引かれており...
                    if border.left is not None and border.left.style == 'thick':
                        # セルの下辺にも太い罫線が引かれていれば、［ラスト・シブリング］だ
                        if border.bottom is not None and border.bottom.style == 'thick':
                            row_th_of_prev_last_underline = -1
                            row_th_of_last_underline = -1
                            print(f"[{datetime.datetime.now()}] 消しゴム {column_letter}列第{row_th}行 ラスト・シブリングなので、最後に見つけた左辺に罫線のないアンダーラインのことは忘れて仕切り直し")
                            shall_break = True

                        # 次行へ読み進めていく
                        else:
                            print(f"[{datetime.datetime.now()}] 消しゴム {column_letter}列第{row_th}行 左側に罫線")
                            pass

                    # セルの左辺に太い罫線が引かれておらず、セルの下辺に太い罫線が引かれていたら、つながっていない垂線だ。それが第何行か覚えておいて仕切り直す
                    elif border.bottom is not None and border.bottom.style == 'thick':
                        row_th_of_prev_last_underline = row_th_of_last_underline
                        row_th_of_last_underline = row_th
                        print(f"[{datetime.datetime.now()}] 消しゴム {column_letter}列第{row_th}行 最後に見つけた、左辺に罫線のないアンダーラインが第何行か覚えておく（第{row_th_of_last_underline}行）（１つ前は第{row_th_of_prev_last_underline}行）")
                        shall_break = True

                    # セルの左辺にも、下辺にも、太い罫線が引かれていなければ、仕切り直し
                    else:
                        shall_break = True
                        print(f"[{datetime.datetime.now()}] 消しゴム {column_letter}列第{row_th}行 セルの左辺にも下辺にも罫線が引かれていなかったので、仕切り直し")


                row_th += 1

                if shall_break:
                    break


            # 消しゴムを掛ける
            start_row_to_erase = row_th_of_prev_last_underline + 1
            end_row_to_erase = row_th_of_last_underline

            if row_th_of_last_underline != -1 and 0 < start_row_to_erase and start_row_to_erase < end_row_to_erase:
                print(f"[{datetime.datetime.now()}] 消しゴム {column_letter}列 消しゴムを掛けたいのは第{start_row_to_erase}～{end_row_to_erase - 1}行")
                for row_th_to_erase in range(start_row_to_erase, end_row_to_erase):
                    # 消すか、見え消しにするか切り替えられるようにしておく
                    ws[f'{column_letter}{row_th_to_erase}'].border = striked_border

        print(f"[{datetime.datetime.now()}] 消しゴム {column_letter}列第{row_th}行 消しゴム掛け終わり（最終は第{ws.max_row}行）")
